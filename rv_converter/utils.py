import json
import os
from typing import Dict, List, Optional

from math import ceil
import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .constants import EXCEL_MAX_COLUMNS, EXCEL_MAX_ROWS
from .i18n import get_i18n

i18n = get_i18n()

try:
    import ijson

    HAS_IJSON = True
except ImportError:
    HAS_IJSON = False


def add_count(ws: Worksheet, num_columns: int, num_rows: int) -> None:
    """Добавляет строку с формулами SUBTOTAL для подсчёта заполненных ячеек.

    Использует SUBTOTAL(103, ...) — аналог COUNTA, который игнорирует
    скрытые (отфильтрованные) строки.

    Args:
        ws: Лист Excel для записи.
        num_columns: Количество столбцов.
        num_rows: Количество строк данных.
    """
    data_start_row = 3  # Row 1 = counts, Row 2 = headers, Row 3+ = data
    data_end_row = data_start_row + num_rows - 1
    formulas = []
    for col_idx in range(1, num_columns + 1):
        col_letter = get_column_letter(col_idx)
        formulas.append(
            f"=SUBTOTAL(103,{col_letter}{data_start_row}:{col_letter}{data_end_row})"
        )
    ws.append(formulas)


def make_bold_and_freeze(ws: Worksheet, row: int) -> None:
    """Делает заголовки жирными и закрепляет строки выше указанной.

    Args:
        ws: Лист Excel для форматирования.
        row: Номер строки, до которой применится форматирование.
    """
    for cell in ws.iter_rows(min_row=1, max_row=row - 1):
        for c in cell:
            c.font = Font(bold=True)

    ws.freeze_panes = ws[f"A{row}"]
    ws.delete_rows(row)


def add_filters(ws: Worksheet, num_columns: int, row: int) -> None:
    """Добавляет автофильтры к указанной строке листа Excel.

    Args:
        ws: Лист Excel для добавления фильтров.
        num_columns: Количество столбцов для фильтрации.
        row: Номер строки с заголовками.
    """
    ws.auto_filter.ref = f"A{row}:{get_column_letter(num_columns)}{row}"


def validate_input_file(input_file_path: str) -> bool:
    """Проверяет существование файла и его расширение (.json или .jsonl).

    Args:
        input_file_path: Путь к файлу для проверки.

    Returns:
        True если файл существует и имеет корректное расширение, иначе False.
    """
    if not os.path.isfile(input_file_path):
        print(i18n.get("file_not_exist", input_file_path))
        return False
    if not (input_file_path.endswith(".json") or input_file_path.endswith(".jsonl")):
        print(i18n.get("file_not_json", input_file_path))
        return False

    return True


def read_json_file(input_file_path: str) -> Optional[List[Dict]]:
    """Читает и парсит JSON-файл. Использует ijson при наличии для потоковой обработки.

    Args:
        input_file_path: Путь к JSON-файлу.

    Returns:
        Список объектов из файла или None при ошибке.
    """
    try:
        with open(input_file_path, "r", encoding="utf-8") as file:
            if HAS_IJSON:
                data = ijson.items(file, "item")
                return list(data)
            else:
                return json.load(file)
    except json.JSONDecodeError:
        print(i18n.get("file_invalid_json", input_file_path))
        return None
    except Exception as e:
        print(i18n.get("file_read_error", str(e)))
        return None


def read_jsonl_file(input_file_path: str) -> Optional[List[Dict]]:
    try:
        data = []
        with open(input_file_path, "r", encoding="utf-8") as file:
            for line in file:
                if not line.strip():
                    continue

                data.append(json.loads(line))
        return data
    except json.JSONDecodeError as e:
        print(i18n.get("file_invalid_jsonl", input_file_path, str(e)))
        return None
    except Exception as e:
        print(i18n.get("file_read_error", str(e)))
        return None


def load_data(filename: str) -> Optional[List[Dict]]:
    if not validate_input_file(filename):
        return None

    if filename.endswith(".jsonl"):
        return read_jsonl_file(filename)
    else:
        return read_json_file(filename)


def split_dataframe(df: pd.DataFrame) -> List[pd.DataFrame]:
    """Разбивает DataFrame на части по максимальному числу строк Excel.

    Args:
        df: DataFrame для разбиения.

    Returns:
        Список DataFrame-частей.
    """
    return [df[i : i + EXCEL_MAX_ROWS] for i in range(0, len(df), EXCEL_MAX_ROWS)]


def split_dataframe_by_columns(df: pd.DataFrame) -> List[pd.DataFrame]:
    """Разбивает DataFrame на части по максимальному числу столбцов Excel.

    Args:
        df: DataFrame для разбиения.

    Returns:
        Список DataFrame-частей.
    """
    num_chunks = ceil(df.shape[1] / EXCEL_MAX_COLUMNS)
    return [
        df.iloc[:, i * EXCEL_MAX_COLUMNS : (i + 1) * EXCEL_MAX_COLUMNS]
        for i in range(num_chunks)
    ]


def split_data(data: List[Dict]) -> List[List[Dict]]:
    """Разбивает список словарей на части по максимальному числу строк Excel.

    Args:
        data: Список словарей для разбиения.

    Returns:
        Список списков словарей.
    """
    return [data[i : i + EXCEL_MAX_ROWS] for i in range(0, len(data), EXCEL_MAX_ROWS)]
